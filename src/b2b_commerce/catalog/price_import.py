import hashlib
import re
import zipfile
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from enum import StrEnum
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

STOCK_LABEL_TO_QUANTITY: dict[str, int] = {
    "Много": 50,
    "Достаточно": 10,
    "Мало": 3,
    "Sold_Out": 0,
}

CATEGORY_ALIASES = {"Рактеки": "Ракетки"}
VALID_CATEGORIES = frozenset({"Мячи", "Аксессуары", "Ракетки"})

ACCESSORY_NAME_BRANDS = (
    "Wilson",
    "Vibora",
    "Elite",
    "Bullpadel",
    "Adidas",
    "Head",
    "Nox",
    "Siux",
)

DATA_NAME_COL = 5
DATA_CARD_PRICE_COL = 9
DATA_SALE_PRICE_COL = 10
DATA_STOCK_COL = 13
DATA_FIRST_PRODUCT_ROW = 7

PRICE_NAME_COL = 3
PRICE_STOCK_COL = 4
PRICE_FIRST_PRODUCT_ROW = 14

DRAWING_NS = {"xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"}
REL_NS = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
OFFICE_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
MODEL_YEAR_RE = re.compile(r"\b(20\d{2})\b")


class ImportRowStatus(StrEnum):
    CREATE = "create"
    UPDATE = "update"
    ERROR = "error"
    SKIP = "skip"


@dataclass(frozen=True)
class ProductImportRow:
    name: str
    brand_name: str
    category_name: str
    model_year: int | None
    sale_price: Decimal
    stock_label: str
    stock_quantity: int
    status_hint: str
    image_bytes: bytes | None
    image_hash: str | None
    source_row: int


@dataclass
class ValidatedImportRow:
    row: ProductImportRow
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


@dataclass
class ImportRowResult:
    source_row: int
    name: str
    brand_name: str
    status: ImportRowStatus
    messages: list[str] = field(default_factory=list)


@dataclass
class ImportReport:
    created: int = 0
    updated: int = 0
    errors: int = 0
    skipped: int = 0
    warnings: int = 0
    row_results: list[ImportRowResult] = field(default_factory=list)


@dataclass(frozen=True)
class _PriceSheetMeta:
    price_row: int
    category_name: str
    brand_sub: str | None


# Нормализует текст для поиска.
def normalize_match_text(value: str) -> str:
    return " ".join(value.strip().lower().split())


# Нормализует текст для категории.
def normalize_category_name(value: str) -> str:
    cleaned = value.strip()
    return CATEGORY_ALIASES.get(cleaned, cleaned)


# Извлекает год модели из названия.
def extract_model_year(name: str) -> int | None:
    match = MODEL_YEAR_RE.search(name)
    if match is None:
        return None
    return int(match.group(1))


# Очищает подстроку бренда.
def clean_brand_sub(value: str) -> str:
    text = value.strip()
    if text.startswith("Fakun"):
        return "Fakun"
    return text


# Определяет имя бренда по названию товара и категории.
def resolve_brand_name(name: str, category_name: str, brand_sub: str | None) -> str:
    if brand_sub:
        return clean_brand_sub(brand_sub)
    if category_name == "Мячи":
        match = re.match(r"Мячи для падел\s+(\S+)", name, flags=re.IGNORECASE)
        if match:
            return match.group(1)
    if category_name == "Аксессуары":
        lowered = name.lower()
        for brand in ACCESSORY_NAME_BRANDS:
            if brand.lower() in lowered:
                return brand
        return "Accessory"
    return "Unknown"


# Преобразует метку наличия в количество.
def map_stock_label(label: str) -> int:
    normalized = label.strip()
    if normalized not in STOCK_LABEL_TO_QUANTITY:
        raise ValueError(f"Неизвестная метка наличия: {label}")
    return STOCK_LABEL_TO_QUANTITY[normalized]


# Преобразует метку наличия в статус.
def status_hint_from_stock(stock_label: str) -> str:
    return "inactive" if stock_label.strip() == "Sold_Out" else "active"


# Преобразует значение в десятичное число.
def _parse_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


# Определяет тип контента изображения по расширению файла.
def _guess_content_type(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    if ext == ".png":
        return "image/png"
    if ext in {".jpg", ".jpeg"}:
        return "image/jpeg"
    return "image/jpeg"


# Загружает изображения из файла.
def _load_price_sheet_images(xlsx_path: Path) -> dict[int, tuple[bytes, str]]:
    with zipfile.ZipFile(xlsx_path) as zf:
        drawing_xml = zf.read("xl/drawings/drawing1.xml")
        rels_xml = zf.read("xl/drawings/_rels/drawing1.xml.rels")

    rels_root = ET.fromstring(rels_xml)
    embed_to_media: dict[str, str] = {}
    for rel in rels_root.findall("rel:Relationship", REL_NS):
        rel_id = rel.get("Id")
        target = rel.get("Target")
        if not rel_id or not target:
            continue
        media_name = Path(target).name
        embed_to_media[rel_id] = media_name

    root = ET.fromstring(drawing_xml)
    images: dict[int, tuple[bytes, str]] = {}
    with zipfile.ZipFile(xlsx_path) as zf:
        for tag in ("twoCellAnchor", "oneCellAnchor"):
            for anchor in root.findall(f".//xdr:{tag}", DRAWING_NS):
                from_el = anchor.find("xdr:from", DRAWING_NS)
                if from_el is None:
                    continue
                row_el = from_el.find("xdr:row", DRAWING_NS)
                if row_el is None or row_el.text is None:
                    continue
                row = int(row_el.text) + 1
                blip = anchor.find(
                    ".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip"
                )
                if blip is None:
                    continue
                embed = blip.get(OFFICE_REL)
                if not embed or embed not in embed_to_media:
                    continue
                media_name = embed_to_media[embed]
                media_path = f"xl/media/{media_name}"
                images[row] = (zf.read(media_path), _guess_content_type(media_name))
    return images


def _parse_price_sheet_meta(price_ws) -> dict[str, _PriceSheetMeta]:
    by_name: dict[str, _PriceSheetMeta] = {}
    category_name: str | None = None
    brand_sub: str | None = None

    for row_idx in range(PRICE_FIRST_PRODUCT_ROW, price_ws.max_row + 1):
        raw_name = price_ws.cell(row_idx, PRICE_NAME_COL).value
        raw_stock = price_ws.cell(row_idx, PRICE_STOCK_COL).value
        if raw_name is None:
            continue
        text = str(raw_name).strip()
        if not text:
            continue

        if text.endswith(":"):
            header = normalize_category_name(text.rstrip(":").strip())
            if header in VALID_CATEGORIES:
                category_name = header
                brand_sub = None
            continue

        if raw_stock in (None, "") and category_name == "Ракетки":
            if not MODEL_YEAR_RE.search(text):
                brand_sub = clean_brand_sub(text)
                continue

        if raw_stock in (None, ""):
            continue

        by_name[text] = _PriceSheetMeta(
            price_row=row_idx,
            category_name=category_name or "",
            brand_sub=brand_sub,
        )
    return by_name


class ExcelParser:

    def __init__(self, path: Path | str) -> None:
        self.path = Path(path)
    # Парсит строки из файла.
    def parse_raw_rows(self) -> list[dict[str, object]]:
        wb = openpyxl.load_workbook(self.path, data_only=True)
        data_ws = wb["data"]
        price_meta = _parse_price_sheet_meta(wb["Прайс"])
        images = _load_price_sheet_images(self.path)

        rows: list[dict[str, object]] = []
        for row_idx in range(DATA_FIRST_PRODUCT_ROW, data_ws.max_row + 1):
            raw_name = data_ws.cell(row_idx, DATA_NAME_COL).value
            if raw_name is None or not str(raw_name).strip():
                continue
            name = str(raw_name).strip()
            meta = price_meta.get(name)
            image_bytes = None
            content_type = None
            if meta is not None and meta.price_row in images:
                image_bytes, content_type = images[meta.price_row]
            rows.append(
                {
                    "source_row": row_idx,
                    "name": name,
                    "card_price": data_ws.cell(row_idx, DATA_CARD_PRICE_COL).value,
                    "sale_price": data_ws.cell(row_idx, DATA_SALE_PRICE_COL).value,
                    "stock_label": data_ws.cell(row_idx, DATA_STOCK_COL).value,
                    "category_name": meta.category_name if meta else "",
                    "brand_sub": meta.brand_sub if meta else None,
                    "image_bytes": image_bytes,
                    "image_content_type": content_type,
                }
            )
        return rows


class Normalizer:
    # Нормализует строку импорта.
    @staticmethod
    def normalize(raw: dict[str, object]) -> ProductImportRow:
        name = str(raw["name"]).strip()
        category_name = normalize_category_name(str(raw.get("category_name") or ""))
        brand_name = resolve_brand_name(name, category_name, raw.get("brand_sub"))
        stock_label = str(raw.get("stock_label") or "").strip()
        sale_price = _parse_decimal(raw.get("sale_price"))
        if sale_price is None:
            sale_price = Decimal("0")
        if stock_label in STOCK_LABEL_TO_QUANTITY:
            stock_quantity = map_stock_label(stock_label)
        else:
            stock_quantity = 0
        image_bytes = raw.get("image_bytes")
        image_hash = None
        if isinstance(image_bytes, (bytes, bytearray)) and image_bytes:
            image_hash = hashlib.sha256(image_bytes).hexdigest()
        return ProductImportRow(
            name=name,
            brand_name=brand_name,
            category_name=category_name,
            model_year=extract_model_year(name),
            sale_price=sale_price,
            stock_label=stock_label,
            stock_quantity=stock_quantity,
            status_hint=status_hint_from_stock(stock_label),
            image_bytes=bytes(image_bytes) if image_bytes else None,
            image_hash=image_hash,
            source_row=int(raw["source_row"]),
        )


class Validator:

    # Проверяет строку импорта.
    @staticmethod
    def validate(row: ProductImportRow, raw: dict[str, object] | None = None) -> ValidatedImportRow:
        result = ValidatedImportRow(row=row)
        if not row.name.strip():
            result.errors.append("Пустое название")
        if row.sale_price <= 0:
            result.errors.append("Некорректная цена ИП")
        if not row.stock_label:
            result.errors.append("Не указана метка наличия")
        elif row.stock_label not in STOCK_LABEL_TO_QUANTITY:
            result.errors.append(f"Неизвестная метка наличия: {row.stock_label}")
        if not row.category_name:
            result.errors.append("Не удалось определить категорию")
        if row.brand_name == "Unknown":
            result.errors.append("Не удалось определить бренд")
        if raw is not None and raw.get("card_price") in (None, ""):
            result.warnings.append("Отсутствует цена по карте")
        if row.image_bytes is None:
            result.warnings.append("Изображение не найдено")
        return result


# Парсит файл Excel с прайсом.
def parse_excel_price_file(path: Path | str) -> tuple[list[ValidatedImportRow], list[str]]:
    parser = ExcelParser(path)
    warnings: list[str] = []
    validated_rows: list[ValidatedImportRow] = []
    for raw in parser.parse_raw_rows():
        row = Normalizer.normalize(raw)
        validated = Validator.validate(row, raw)
        warnings.extend(validated.warnings)
        validated_rows.append(validated)
    return validated_rows, warnings
